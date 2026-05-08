"""
Universal machine scraper.
Usage:  python run.py <website_url>
Example: python run.py https://valmatrading.com
         python run.py https://exapro.com/en/machine-tools/
"""

import sys
import os
import re
import time
import json
import requests
from bs4 import BeautifulSoup
from datetime import datetime
from urllib.parse import urlparse, urljoin
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_DIR = "output"
NA         = "Not Available"
HEADERS    = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
    )
}

# ── Shared helpers ────────────────────────────────────────────────────────────

def get_soup(url):
    for attempt in range(3):
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            r.raise_for_status()
            r.encoding = r.apparent_encoding or "utf-8"
            return BeautifulSoup(r.text, "lxml")
        except Exception as e:
            if attempt == 2:
                print(f"    [ERROR] {url} -> {e}")
            time.sleep(2)
    return None


def clean(text):
    if not text:
        return ""
    text = text.replace("\xa0", " ").replace("’", "'")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def save_excel(records, domain):
    if not records:
        print("No data to save.")
        return None

    df = pd.DataFrame(records)
    for col in ["Type", "Brand", "Model", "Description", "Price", "Image URL", "Source URL"]:
        if col not in df.columns:
            df[col] = ""
    df = df[["Type", "Brand", "Model", "Description", "Price", "Image URL", "Source URL"]]

    df["Brand"].fillna(NA, inplace=True)
    df["Model"].fillna(NA, inplace=True)
    df["Description"].fillna(NA, inplace=True)
    df["Price"].fillna("", inplace=True)
    df["Image URL"].fillna("", inplace=True)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    site_name = domain.replace("www.", "").replace(".", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filepath  = os.path.join(OUTPUT_DIR, f"{site_name}_{timestamp}.xlsx")

    df.to_excel(filepath, index=False, engine="openpyxl")

    wb = load_workbook(filepath)
    ws = wb.active

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    ws.freeze_panes = "A2"

    for col_idx, col_cells in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 80)

    wb.save(filepath)
    return filepath


# ── Known-site runners ────────────────────────────────────────────────────────

def run_bg_used(_url):
    import scraper as bg
    bg.main()


def run_valma(_url):
    import scraper_valma as valma
    valma.main()


def run_ajmera(_url):
    import scraper_ajmera as ajmera
    ajmera.main()


# ── Generic scraper ───────────────────────────────────────────────────────────

# URL path patterns that suggest a product detail page
PRODUCT_HINTS = [
    r"/product[s]?/[^/]+/?$",
    r"/machine[s]?/[^/]+/?$",
    r"/item[s]?/[^/]+/?$",
    r"/listing[s]?/[^/]+/?$",
    r"/equipment/[^/]+/?$",
    r"/used-[^/]+/?$",
    r"/second-hand/[^/]+/?$",
    r"/ad/[^/]+/?$",
    r"/p/[^/]+/?$",
    r"[?&]id=\d+",
    r"[?&]ref=",
    r"/indexd",          # bg-used-industry style
    r"/detail",
]

# URL patterns to always skip
SKIP_HINTS = [
    r"/(cart|checkout|login|register|account|contact|about|faq|blog|news|tag)/?(\?|$|#)",
    r"\.(pdf|zip|jpg|jpeg|png|gif|css|js|xml|ico)(\?|$)",
    r"mailto:",
    r"javascript:",
    r"tel:",
]


def looks_like_product(href):
    h = href.lower()
    if any(re.search(p, h) for p in SKIP_HINTS):
        return False
    return any(re.search(p, h) for p in PRODUCT_HINTS)


def same_domain_links(soup, base_url, domain):
    seen, links = set(), []
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if not href or href.startswith("#"):
            continue
        full = urljoin(base_url, href).split("#")[0]
        p = urlparse(full)
        if p.netloc and domain not in p.netloc:
            continue
        if full not in seen:
            seen.add(full)
            links.append(full)
    return links


def collect_product_urls(start_url, domain, base_url):
    print(f"  Scanning start page: {start_url}")
    soup = get_soup(start_url)
    if not soup:
        return []

    all_links   = same_domain_links(soup, base_url, domain)
    product_set = set()
    cat_links   = []

    for link in all_links:
        if looks_like_product(link):
            product_set.add(link)
        elif link != start_url and link.startswith(base_url):
            cat_links.append(link)

    # If not enough direct product links, crawl one level deeper
    if len(product_set) < 5:
        print(f"  Crawling {min(len(cat_links), 40)} sub-pages...")
        crawled = {start_url}
        for cat_url in cat_links[:40]:
            if cat_url in crawled:
                continue
            crawled.add(cat_url)
            cat_soup = get_soup(cat_url)
            if not cat_soup:
                continue
            for link in same_domain_links(cat_soup, base_url, domain):
                if looks_like_product(link):
                    product_set.add(link)
            time.sleep(0.3)

    print(f"  Total product URLs found: {len(product_set)}")
    return list(product_set)


def extract_json_ld(soup):
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            if isinstance(data, list):
                data = data[0] if data else {}
            t = data.get("@type", "")
            if any(k in t for k in ["Product", "Machine", "Item", "Offer", "Vehicle"]):
                return data
        except Exception:
            pass
    return {}


def extract_og(soup):
    og = {}
    for meta in soup.find_all("meta"):
        prop    = meta.get("property", "") or meta.get("name", "")
        content = meta.get("content", "")
        if prop.startswith("og:") and content:
            og[prop[3:]] = content
    return og


def labeled_value(lines, labels):
    for i, line in enumerate(lines):
        low = line.lower().rstrip(":")
        for lbl in labels:
            if low == lbl.lower():
                for j in range(i + 1, min(i + 4, len(lines))):
                    cand = lines[j]
                    if cand and cand.lower() not in [l.lower() for l in labels]:
                        return cand
            if line.lower().startswith(lbl.lower() + ":"):
                val = line[len(lbl) + 1:].strip()
                if val:
                    return val
    return NA


def generic_parse_page(url):
    soup = get_soup(url)
    if not soup:
        return None

    for tag in soup(["script", "style", "nav", "footer", "header"]):
        tag.decompose()

    schema = extract_json_ld(soup)
    og     = extract_og(soup)
    main   = soup.find("main") or soup.find("article") or soup.body
    lines  = [clean(l) for l in main.get_text(separator="\n").splitlines() if clean(l)]

    # ── Description ──────────────────────────────────────────────────────────
    description = NA
    if schema.get("name"):
        description = clean(schema["name"])
    elif og.get("title"):
        description = clean(og["title"]).split("|")[0].strip()
    else:
        h1 = soup.find("h1")
        if h1:
            description = clean(h1.get_text())
        else:
            title_tag = soup.find("title")
            if title_tag:
                description = clean(title_tag.get_text()).split("|")[0].strip()

    # ── Brand ─────────────────────────────────────────────────────────────────
    brand = NA
    if schema.get("brand"):
        b = schema["brand"]
        brand = clean(b.get("name", b) if isinstance(b, dict) else str(b))
    if brand == NA:
        brand = labeled_value(lines, ["Brand", "Manufacturer", "Make", "Marque", "Hersteller", "Fabricant"])
    if brand == NA and description != NA:
        words = description.split()
        if words:
            brand = words[0].strip(",:;-()")

    # ── Model ─────────────────────────────────────────────────────────────────
    model = NA
    if schema.get("model"):
        model = clean(str(schema["model"]))
    if model == NA:
        model = labeled_value(lines, ["Model", "Model number", "Type", "Ref", "Reference", "Modèle"])

    # ── Price ─────────────────────────────────────────────────────────────────
    price = ""
    if schema.get("offers"):
        offers = schema["offers"]
        if isinstance(offers, list):
            offers = offers[0]
        p        = offers.get("price", "")
        currency = offers.get("priceCurrency", "")
        if p:
            price = f"{currency} {p}".strip() if currency else str(p)

    if not price:
        val = labeled_value(lines, ["Price", "Asking price", "Selling price", "Prix", "Preis"])
        if val != NA:
            price = val

    if not price:
        for line in lines:
            if re.search(r"(€|£|\$|EUR|USD|GBP)\s?[\d,. ]+", line):
                m = re.search(r"(€|£|\$|EUR|USD|GBP)\s?[\d,. ]+", line)
                price = m.group(0).strip()
                break

    # ── Image URL ─────────────────────────────────────────────────────────────
    image_url = ""
    if schema.get("image"):
        img = schema["image"]
        if isinstance(img, list):
            img = img[0]
        image_url = img.get("url", img) if isinstance(img, dict) else str(img)
    elif og.get("image"):
        image_url = og["image"]

    if not image_url:
        for img in soup.find_all("img", src=True):
            src = img.get("src", "")
            if not src:
                continue
            if any(s in src.lower() for s in ["logo", "icon", "banner", "flag", "avatar", "placeholder", "sprite"]):
                continue
            try:
                if int(str(img.get("width", "200")).replace("px", "")) < 80:
                    continue
            except Exception:
                pass
            full_src = urljoin(url, src)
            full_src = re.sub(r"-\d+x\d+(\.(jpg|jpeg|png|webp))$", r"\1", full_src, flags=re.IGNORECASE)
            full_src = re.sub(r"_m(\.(jpg|jpeg|png))$",            r"\1", full_src, flags=re.IGNORECASE)
            image_url = full_src
            break

    # ── Type (from breadcrumb or URL path) ───────────────────────────────────
    machine_type = NA
    breadcrumb = soup.find(["nav", "div", "ol", "ul"], class_=re.compile(r"breadcrumb", re.I))
    if breadcrumb:
        crumbs = [clean(a.get_text()) for a in breadcrumb.find_all("a") if clean(a.get_text())]
        if len(crumbs) >= 2:
            machine_type = crumbs[-1]

    if machine_type == NA:
        path  = urlparse(url).path
        parts = [p for p in path.strip("/").split("/")
                 if p and p not in {"products", "machines", "items", "used",
                                    "second-hand", "en", "de", "fr", "listing"}]
        if parts:
            machine_type = parts[0].replace("-", " ").replace("_", " ").title()

    if machine_type == NA:
        machine_type = urlparse(url).netloc.replace("www.", "")

    return {
        "Type":        machine_type,
        "Brand":       brand,
        "Model":       model,
        "Description": description,
        "Price":       price,
        "Image URL":   image_url,
        "Source URL":  url,
    }


def run_generic(start_url):
    parsed   = urlparse(start_url)
    domain   = parsed.netloc
    base_url = f"{parsed.scheme}://{domain}"

    print(f"Auto scraper -> {start_url}")
    print("=" * 55)

    print("Step 1: Finding product pages...")
    product_urls = collect_product_urls(start_url, domain, base_url)

    if not product_urls:
        print("\nNo product pages found.")
        print("Tip: Try passing a specific category page URL instead of the homepage.")
        print("     Example: python run.py https://example.com/used-machines/cnc-lathes/")
        return

    print(f"\nStep 2: Scraping {len(product_urls)} pages...")
    records = []
    for i, url in enumerate(product_urls, 1):
        print(f"  [{i}/{len(product_urls)}] {url}")
        try:
            data = generic_parse_page(url)
            if data:
                records.append(data)
        except Exception as e:
            print(f"    [SKIP] {e}")
        time.sleep(0.5)

    if not records:
        print("No data extracted.")
        return

    filepath = save_excel(records, domain)

    print("\n" + "=" * 55)
    print("  CRAWL SUMMARY")
    print("=" * 55)
    print(f"  Total machines  : {len(records)}")
    print(f"  Excel saved at  : {os.path.abspath(filepath)}")
    print("=" * 55)
    print(f"\nDone. Open: {os.path.abspath(filepath)}")


# ── Dispatcher ────────────────────────────────────────────────────────────────

KNOWN_SITES = {
    "bg-used-industry.com": run_bg_used,
    "valmatrading.com":     run_valma,
    "ajmeramachines.com":   run_ajmera,
}


def main():
    if len(sys.argv) < 2:
        print("=" * 55)
        print("  Machine Scraper — Universal")
        print("=" * 55)
        print("  Usage:   python run.py <website_url>")
        print("  Example: python run.py https://valmatrading.com")
        print("           python run.py https://exapro.com/en/")
        print("=" * 55)
        sys.exit(1)

    url = sys.argv[1].strip()
    if not url.startswith("http"):
        url = "https://" + url

    domain = urlparse(url).netloc.replace("www.", "")

    for known_domain, handler in KNOWN_SITES.items():
        if known_domain in domain:
            print(f"Known site: {known_domain} — using optimized scraper")
            print("=" * 55)
            handler(url)
            return

    run_generic(url)


if __name__ == "__main__":
    main()
