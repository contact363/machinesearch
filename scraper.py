import os
import re
import time
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
BASE_URL   = "https://www.bg-used-industry.com"
OUTPUT_DIR = "output"
NA         = "Not Available"
HEADERS    = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
    )
}
SKIP_BRAND_WORDS = {"WANTED", "FOR", "SALE", "URGENT", "NEEDED"}

# ── TARGET CATEGORIES ─────────────────────────
# Format: "User-friendly Type Name": [list of subcategory URLs]
TARGETS = {
    "VMC": [
        "https://www.bg-used-industry.com/?t_c=33&t=33&lng=en",   # Vertical
    ],
    "HMC": [
        "https://www.bg-used-industry.com/?t_c=33&t=34&lng=en",   # Horizontal
        "https://www.bg-used-industry.com/?t_c=33&t=177&lng=en",  # Universal
    ],
    "CNC Lathe": [
        "https://www.bg-used-industry.com/?t_c=66&t=79&lng=en",   # CNC Fixed Head <2m DBC
        "https://www.bg-used-industry.com/?t_c=66&t=80&lng=en",   # CNC Fixed Head >=2m DBC
        "https://www.bg-used-industry.com/?t_c=66&t=171&lng=en",  # CNC Sliding Head
        "https://www.bg-used-industry.com/?t_c=66&t=75&lng=en",   # Turning Centres
    ],
    "Cylindrical Grinder": [
        "https://www.bg-used-industry.com/?t_c=106&t=117&lng=en", # Cylindrical Plain
        "https://www.bg-used-industry.com/?t_c=106&t=119&lng=en", # Cylindrical Plain CNC
        "https://www.bg-used-industry.com/?t_c=106&t=118&lng=en", # Cylindrical Universal
        "https://www.bg-used-industry.com/?t_c=106&t=584&lng=en", # Cylindrical Universal CNC
        "https://www.bg-used-industry.com/?t_c=106&t=108&lng=en", # Cylindrical Internal
        "https://www.bg-used-industry.com/?t_c=106&t=109&lng=en", # Cylindrical Internal CNC
    ],
    "Injection Molding Machine": [
        "https://www.bg-used-industry.com/?t_c=155&t=636&lng=en", # Plastics Machinery
        "https://www.bg-used-industry.com/?t_c=3&t=3&lng=en",     # Forming Machinery
    ],
    "Gear Grinder": [
        "https://www.bg-used-industry.com/?t_c=6&t=13&lng=en",    # Gear Machinery Grinders
        "https://www.bg-used-industry.com/?t_c=6&t=598&lng=en",   # Gear Machinery Grinders CNC
    ],
    "Surface Grinder": [
        "https://www.bg-used-industry.com/?t_c=106&t=114&lng=en", # Surface Reciprocating Table
        "https://www.bg-used-industry.com/?t_c=106&t=637&lng=en", # Surface Reciprocating Table CNC
        "https://www.bg-used-industry.com/?t_c=106&t=629&lng=en", # Surface with vertical spindle
    ],
    "Centerless Grinder": [
        "https://www.bg-used-industry.com/?t_c=106&t=106&lng=en", # Centreless
        "https://www.bg-used-industry.com/?t_c=106&t=631&lng=en", # Centreless CNC
    ],
}
# ─────────────────────────────────────────────


def get_soup(url):
    for attempt in range(3):
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            r.raise_for_status()
            r.encoding = "utf-8"
            return BeautifulSoup(r.text, "lxml")
        except Exception as e:
            if attempt == 2:
                print(f"    [ERROR] {url} -> {e}")
            time.sleep(2)
    return None


def clean(text):
    if not text:
        return ""
    text = text.replace("\xa0", " ").replace("’", "'")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def first_word(text):
    for word in text.split():
        w = word.strip(",:;-()").upper()
        if w and w not in SKIP_BRAND_WORDS:
            return word.strip(",:;-()")
    return NA


def before_comma(text):
    if "," in text:
        return text.split(",")[0].strip()
    return text.strip()


def collect_machine_urls(listing_url, machine_type):
    machine_urls = []
    seen = set()
    url  = listing_url

    while url:
        soup = get_soup(url)
        if not soup:
            break

        for a in soup.find_all("a", href=True):
            href = a["href"]
            if "indexd" not in href or "id=" not in href:
                continue
            full = href if href.startswith("http") else f"{BASE_URL}/{href.lstrip('/')}"
            if "lng=en" not in full:
                full += "&lng=en"
            if full not in seen:
                seen.add(full)
                machine_urls.append({"url": full, "type": machine_type})

        # Pagination
        next_url   = None
        current_pg = 0
        if "pageNum_Recordset1=" in url:
            try:
                current_pg = int(url.split("pageNum_Recordset1=")[1].split("&")[0])
            except Exception:
                pass

        for a in soup.find_all("a", href=True):
            href = a["href"]
            if "pageNum_Recordset1=" not in href:
                continue
            try:
                pg = int(href.split("pageNum_Recordset1=")[1].split("&")[0])
            except Exception:
                continue
            if pg > current_pg:
                next_url = href if href.startswith("http") else f"{BASE_URL}/{href.lstrip('/')}"
                break

        url = next_url
        if next_url:
            time.sleep(0.4)

    return machine_urls


def parse_machine(url, machine_type):
    soup = get_soup(url)
    if not soup:
        return None

    tdprd = soup.find("td", class_="TdPrd")

    # Description: td.head3
    head3       = soup.find("td", class_="head3")
    description = clean(head3.get_text()) if head3 else NA

    # Brand: first meaningful word of description
    brand = first_word(description) if description != NA else NA

    # Model: first td.body2 inside TdPrd (text before first comma, max 35 chars)
    model = NA
    if tdprd:
        body2 = tdprd.find("td", class_="body2")
        if body2:
            raw = clean(body2.get_text())
            if raw:
                candidate = before_comma(raw)
                if candidate and len(candidate) <= 35:
                    model = candidate

    # Price: td.price — leave empty string if not available
    price_td = soup.find("td", class_="price")
    if price_td:
        price = clean(price_td.get_text())
    else:
        price = ""
        for td in soup.find_all("td", class_="body1"):
            if "poa" in clean(td.get_text()).lower():
                price = "POA"
                break

    # Image URL: first photos/ img — use full size (remove _m suffix)
    image_url = ""
    for img in soup.find_all("img", src=True):
        src = img["src"]
        if "photos/" in src:
            # Convert thumbnail (_m.jpg / _m.jpeg) to full size
            src = re.sub(r"_m(\.(jpg|jpeg|png))$", r"\1", src, flags=re.IGNORECASE)
            image_url = src if src.startswith("http") else f"{BASE_URL}/{src.lstrip('/')}"
            break

    missing = [f for f, v in {"Brand": brand, "Model": model}.items() if v == NA]
    if missing:
        print(f"    [MISSING] {', '.join(missing)}")

    return {
        "Type":        machine_type,
        "Brand":       brand,
        "Model":       model,
        "Description": description,
        "Price":       price,
        "Image URL":   image_url,
        "Source URL":  url,
    }


def save_excel(df, filepath):
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    df.to_excel(filepath, index=False, engine="openpyxl")

    wb = load_workbook(filepath)
    ws = wb.active

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    ws.freeze_panes = "A2"

    for col_idx, col_cells in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len    = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 80)

    wb.save(filepath)


def main():
    print("Targeted scraper — only scraping selected machine types")
    print("=" * 55)
    for t in TARGETS:
        print(f"  - {t}")
    print("=" * 55)

    # Collect machine URLs for each target type
    print("\nStep 1: Collecting machine URLs...")
    all_refs = []
    seen_urls = set()

    for machine_type, listing_urls in TARGETS.items():
        type_count = 0
        for listing_url in listing_urls:
            refs = collect_machine_urls(listing_url, machine_type)
            for ref in refs:
                if ref["url"] not in seen_urls:
                    seen_urls.add(ref["url"])
                    all_refs.append(ref)
                    type_count += 1
        print(f"  {machine_type:<30}: {type_count} machines")
        time.sleep(0.3)

    print(f"\n  Total unique machines to scrape: {len(all_refs)}")

    # Scrape each machine detail page
    print("\nStep 2: Scraping machine detail pages...")
    all_machines = []
    missing_stats = {"Brand": 0, "Model": 0, "Price": 0, "Image URL": 0}

    for i, ref in enumerate(all_refs, 1):
        print(f"  [{i}/{len(all_refs)}] {ref['url']}")
        try:
            data = parse_machine(ref["url"], ref["type"])
            if data:
                for field in missing_stats:
                    if data.get(field) == NA:
                        missing_stats[field] += 1
                all_machines.append(data)
        except Exception as e:
            print(f"    [SKIP] {e}")
        time.sleep(0.5)

    if not all_machines:
        print("No data collected.")
        return

    df = pd.DataFrame(all_machines)
    total_raw = len(df)
    df.drop_duplicates(subset=["Source URL"], inplace=True)
    dupes = total_raw - len(df)
    # Only fill Brand/Model/Description with NA — leave Price and Image URL empty
    df["Brand"].fillna(NA, inplace=True)
    df["Model"].fillna(NA, inplace=True)
    df["Description"].fillna(NA, inplace=True)
    df["Price"].fillna("", inplace=True)
    df["Image URL"].fillna("", inplace=True)

    # Sort by Type for cleaner Excel
    df.sort_values("Type", inplace=True)
    df.reset_index(drop=True, inplace=True)

    timestamp   = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = os.path.join(OUTPUT_DIR, f"machines_output_{timestamp}.xlsx")
    save_excel(df, output_path)

    print("\n" + "=" * 55)
    print("  CRAWL SUMMARY")
    print("=" * 55)
    print(f"  Total machines scraped   : {total_raw}")
    print(f"  Duplicates removed       : {dupes}")
    print(f"  Final machines in Excel  : {len(df)}")
    print(f"  Excel saved at           : {os.path.abspath(output_path)}")
    print("-" * 55)
    print("  By machine type:")
    for t, cnt in df["Type"].value_counts().items():
        print(f"    {t:<30}: {cnt}")
    print("-" * 55)
    print("  Missing field statistics:")
    for field, count in missing_stats.items():
        pct = (count / total_raw * 100) if total_raw > 0 else 0
        print(f"    {field:<15}: {count:>4} missing ({pct:.1f}%)")
    print("=" * 55)
    print(f"\nDone. Open: {os.path.abspath(output_path)}")


if __name__ == "__main__":
    main()
