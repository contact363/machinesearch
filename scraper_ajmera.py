"""
Scraper for ajmeramachines.com
- Reads all categories from /stocklist
- Reads listing table on /viewall?list=<Category>
- Extracts: Type, Brand, Model, Description, Price, Image URL, Source URL
"""

import os
import re
import time
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_URL   = "https://ajmeramachines.com"
OUTPUT_DIR = "output"
NA         = "Not Available"
HEADERS    = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
    )
}


def get_soup(url):
    for attempt in range(3):
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            r.raise_for_status()
            r.encoding = r.apparent_encoding or "utf-8"
            return BeautifulSoup(r.text, "lxml")
        except Exception as e:
            if attempt == 2:
                print(f"    [ERROR] {url} -> {e}")
            time.sleep(2)
    return None


def clean(text):
    if not text:
        return ""
    text = text.replace("\xa0", " ").replace("'", "'")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def get_categories():
    """Scrape /stocklist and return list of (category_name, viewall_url)."""
    soup = get_soup(f"{BASE_URL}/stocklist")
    if not soup:
        return []

    categories = []
    seen = set()
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if "viewall" in href and "list=" in href:
            full = href if href.startswith("http") else f"{BASE_URL}/{href.lstrip('/')}"
            cat_name = re.search(r"list=(.+?)(&|$)", full)
            if cat_name:
                name = cat_name.group(1).replace("+", " ").strip()
                if name not in seen:
                    seen.add(name)
                    categories.append((name, full))

    return categories


def scrape_category(category_name, listing_url):
    """
    Fetch the category listing page and extract all machines from the table.
    Table columns: Stock#, Image, Make, Model, MYear
    """
    soup = get_soup(listing_url)
    if not soup:
        return []

    machines = []
    rows = soup.find_all("tr")

    for row in rows:
        cols = row.find_all("td")
        if len(cols) < 4:
            continue

        # Find the machine ID from any href in the row
        machine_id = None
        for a in row.find_all("a", href=True):
            m = re.search(r"viewmac\?id=(\d+)", a["href"])
            if m:
                machine_id = m.group(1)
                break

        if not machine_id:
            continue

        # Column 2: Make/Brand (3rd td, index 2)
        brand_td = cols[2] if len(cols) > 2 else None
        brand = clean(brand_td.get_text()) if brand_td else NA
        if not brand:
            brand = NA

        # Column 3: Model (4th td, index 3)
        model_td = cols[3] if len(cols) > 3 else None
        model = clean(model_td.get_text()) if model_td else NA
        if not model:
            model = NA

        # Description = Brand + Model
        description = f"{brand} {model}".strip() if brand != NA or model != NA else NA

        # Image URL — full size: machines/Images/<id>_1.jpg
        image_url = f"{BASE_URL}/machines/Images/{machine_id}_1.jpg"

        # Source URL
        source_url = f"{BASE_URL}/viewmac?id={machine_id}"

        machines.append({
            "Type":        category_name,
            "Brand":       brand,
            "Model":       model,
            "Description": description,
            "Price":       "",
            "Image URL":   image_url,
            "Source URL":  source_url,
        })

    return machines


def save_excel(df, filepath):
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    df.to_excel(filepath, index=False, engine="openpyxl")

    wb = load_workbook(filepath)
    ws = wb.active

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    ws.freeze_panes = "A2"

    for col_idx, col_cells in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    wb.save(filepath)


def main():
    print("Scraper — ajmeramachines.com")
    print("=" * 55)

    # Step 1: Get all categories
    print("Step 1: Reading stocklist categories...")
    categories = get_categories()
    print(f"  Found {len(categories)} categories")
    time.sleep(0.5)

    if not categories:
        print("No categories found. Exiting.")
        return

    # Step 2: Scrape each category listing page
    print("\nStep 2: Scraping machine listings...")
    all_machines = []
    seen_urls    = set()

    for cat_name, cat_url in categories:
        machines = scrape_category(cat_name, cat_url)
        new_count = 0
        for m in machines:
            if m["Source URL"] not in seen_urls:
                seen_urls.add(m["Source URL"])
                all_machines.append(m)
                new_count += 1
        print(f"  {cat_name:<40}: {new_count} machines")
        time.sleep(0.4)

    if not all_machines:
        print("No data collected.")
        return

    df = pd.DataFrame(all_machines)
    total_raw = len(df)
    df.drop_duplicates(subset=["Source URL"], inplace=True)

    df["Brand"].fillna(NA, inplace=True)
    df["Model"].fillna(NA, inplace=True)
    df["Description"].fillna(NA, inplace=True)
    df["Price"].fillna("", inplace=True)
    df["Image URL"].fillna("", inplace=True)

    df.sort_values("Type", inplace=True)
    df.reset_index(drop=True, inplace=True)

    timestamp   = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = os.path.join(OUTPUT_DIR, f"ajmera_output_{timestamp}.xlsx")
    save_excel(df, output_path)

    print("\n" + "=" * 55)
    print("  CRAWL SUMMARY")
    print("=" * 55)
    print(f"  Total machines scraped   : {total_raw}")
    print(f"  Duplicates removed       : {total_raw - len(df)}")
    print(f"  Final machines in Excel  : {len(df)}")
    print(f"  Excel saved at           : {os.path.abspath(output_path)}")
    print("-" * 55)
    print("  By machine type:")
    for t, cnt in df["Type"].value_counts().items():
        t_str = str(t)[:40]
        print(f"    {t_str:<40}: {cnt}")
    print("=" * 55)
    print(f"\nDone. Open: {os.path.abspath(output_path)}")


if __name__ == "__main__":
    main()
