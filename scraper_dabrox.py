"""
Scraper for dabrox.com
- Accepts a category listing URL (e.g. /forging-stamping-machines/screw-presses/)
- Handles pagination automatically (/page/2/, /page/3/ ...)
- Extracts Type, Brand, Model from listing cards + detail page specs table
- Full-size image by stripping size suffix from thumbnail URL
- No price (site uses inquiry form)
"""

import os
import re
import sys
import time
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from urllib.parse import urljoin, urlparse

BASE_URL   = "https://dabrox.com"
OUTPUT_DIR = "output"
NA         = "Not Available"
HEADERS    = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
    )
}


def get_soup(url):
    for attempt in range(3):
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            r.raise_for_status()
            r.encoding = r.apparent_encoding or "utf-8"
            return BeautifulSoup(r.text, "lxml")
        except Exception as e:
            if attempt == 2:
                print(f"    [ERROR] {url} -> {e}")
            time.sleep(2)
    return None


def clean(text):
    if not text:
        return ""
    text = text.replace("\xa0", " ").replace("’", "'").replace("—", "-")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def full_image(src):
    """Remove WordPress-style size suffix e.g. -300x225 before extension."""
    return re.sub(r"-\d+x\d+(\.(jpg|jpeg|png|webp))$", r"\1", src, flags=re.IGNORECASE)


def collect_listing_urls(start_url):
    """Walk all paginated listing pages and return machine detail URLs."""
    detail_urls = []
    seen        = set()
    page_url    = start_url

    while page_url:
        print(f"  Listing page: {page_url}")
        soup = get_soup(page_url)
        if not soup:
            break

        # Non-machine pages to always skip
        SKIP_SLUGS = {
            "about-us", "contact-us", "contact", "about", "privacy-policy",
            "terms", "faq", "search", "blog", "news", "sitemap",
            "login", "register", "cart", "checkout",
        }

        found = 0
        for a in soup.find_all("a", href=True):
            href = a["href"]
            full = href if href.startswith("http") else urljoin(BASE_URL, href)
            parsed = urlparse(full)
            path   = parsed.path.strip("/")
            slug   = path.split("/")[-1] if path else ""
            if (
                "dabrox.com" in parsed.netloc
                and path
                and "/" not in path                          # single slug = detail page
                and full not in seen
                and slug not in SKIP_SLUGS
                and not any(s in full for s in ["/page/", "/category/", "/tag/",
                                                "/forging-stamping-machines/",
                                                "?", "#", ".com/en", ".com/de"])
            ):
                seen.add(full)
                detail_urls.append(full)
                found += 1

        print(f"    Found {found} machines")

        # Next page: go sequentially (current + 1), not jump to highest
        next_url  = None
        current_n = 1
        m = re.search(r"/page/(\d+)/?$", page_url.rstrip("/"))
        if m:
            current_n = int(m.group(1))

        next_n = current_n + 1
        candidate = re.sub(r"/page/\d+/?$", "", page_url.rstrip("/")) + f"/page/{next_n}/"
        # Check if next page exists by looking for its link on current page
        page_nums = set()
        for a in soup.find_all("a", href=True):
            pm = re.search(r"/page/(\d+)/?", a["href"])
            if pm:
                page_nums.add(int(pm.group(1)))
        if next_n in page_nums:
            next_url = candidate
        elif page_nums and max(page_nums) > current_n:
            # Some pages skipped in nav — try next_n directly
            next_url = candidate

        page_url = next_url
        if page_url:
            time.sleep(0.5)

    return detail_urls


def parse_detail(url, category_type):
    """Fetch a machine detail page and extract all fields from the specs table."""
    soup = get_soup(url)
    if not soup:
        return None

    # ── Description from h1 ───────────────────────────────────────────────────
    h1 = soup.find("h1")
    description = clean(h1.get_text()) if h1 else NA

    # ── Specifications table ──────────────────────────────────────────────────
    specs = {}
    for table in soup.find_all("table"):
        for row in table.find_all("tr"):
            cols = row.find_all(["td", "th"])
            if len(cols) >= 2:
                key = clean(cols[0].get_text()).rstrip(":").lower()
                val = clean(cols[1].get_text())
                if key and val:
                    specs[key] = val

    # Also scan <dl> / definition lists common on dabrox
    for dl in soup.find_all("dl"):
        keys   = [clean(dt.get_text()).rstrip(":").lower() for dt in dl.find_all("dt")]
        values = [clean(dd.get_text()) for dd in dl.find_all("dd")]
        for k, v in zip(keys, values):
            if k and v:
                specs[k] = v

    # Also scan labeled <div> / <p> pairs (label: value on same line)
    main = soup.find("main") or soup.find("article") or soup.body
    if main:
        for tag in main(["script", "style", "nav", "footer"]):
            tag.decompose()
        lines = [clean(l) for l in main.get_text(separator="\n").splitlines() if clean(l)]
        for line in lines:
            if ":" in line:
                parts = line.split(":", 1)
                k = parts[0].strip().lower()
                v = parts[1].strip()
                if k and v and len(k) < 40:
                    specs.setdefault(k, v)

    def pick(keys):
        for k in keys:
            for sk, sv in specs.items():
                if k.lower() in sk:
                    return sv
        return NA

    machine_type = pick(["type", "category", "press type", "equipment type"])
    if machine_type == NA:
        machine_type = category_type   # fallback to URL category

    brand = pick(["manufacturer", "brand", "make", "producer"])
    model = pick(["model"])

    # Fallback: parse from h1 — "[Type] [Brand] [Model] — [Capacity]"
    if (brand == NA or model == NA) and description != NA:
        # Remove capacity part after " — " or " - "
        base = re.split(r"\s+[—–-]\s+", description)[0].strip()
        words = base.split()
        if len(words) >= 3 and brand == NA:
            brand = words[1]
        if len(words) >= 3 and model == NA:
            model = " ".join(words[2:])

    # ── Full-size image ───────────────────────────────────────────────────────
    image_url = ""
    for img in soup.find_all("img", src=True):
        src = img["src"]
        if any(s in src.lower() for s in ["logo", "icon", "flag", "avatar", "sprite", "manufacturer"]):
            continue
        if "uploads" in src or "media" in src or "images" in src:
            image_url = full_image(src if src.startswith("http") else urljoin(BASE_URL, src))
            break

    return {
        "Type":        machine_type,
        "Brand":       brand,
        "Model":       model,
        "Description": description,
        "Price":       "",
        "Image URL":   image_url,
        "Source URL":  url,
    }


def save_excel(df, filepath):
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    df.to_excel(filepath, index=False, engine="openpyxl")

    wb = load_workbook(filepath)
    ws = wb.active

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    ws.freeze_panes = "A2"

    for col_idx, col_cells in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len    = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 80)

    wb.save(filepath)


def scrape(start_url):
    # Derive category name from URL path
    path_parts    = [p for p in urlparse(start_url).path.strip("/").split("/") if p]
    category_slug = path_parts[-1] if path_parts else "machines"
    category_type = category_slug.replace("-", " ").title()

    site_label = "dabrox"
    print(f"Scraper — dabrox.com | Category: {category_type}")
    print("=" * 55)

    # Step 1: Collect all detail page URLs from listing
    print("Step 1: Collecting machine URLs from listing pages...")
    detail_urls = collect_listing_urls(start_url)
    print(f"\n  Total machines found: {len(detail_urls)}")

    if not detail_urls:
        print("No machines found. Exiting.")
        return

    # Step 2: Scrape each detail page
    print("\nStep 2: Scraping machine detail pages...")
    all_machines = []

    for i, url in enumerate(detail_urls, 1):
        print(f"  [{i}/{len(detail_urls)}] {url}")
        try:
            data = parse_detail(url, category_type)
            if data:
                all_machines.append(data)
        except Exception as e:
            print(f"    [SKIP] {e}")
        time.sleep(0.5)

    if not all_machines:
        print("No data collected.")
        return

    df = pd.DataFrame(all_machines)
    df = df[["Type", "Brand", "Model", "Description", "Price", "Image URL", "Source URL"]]

    df["Brand"]       = df["Brand"].fillna(NA)
    df["Model"]       = df["Model"].fillna(NA)
    df["Description"] = df["Description"].fillna(NA)
    df["Price"]       = df["Price"].fillna("")
    df["Image URL"]   = df["Image URL"].fillna("")

    df.sort_values("Type", inplace=True)
    df.reset_index(drop=True, inplace=True)

    timestamp   = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = os.path.join(OUTPUT_DIR, f"{site_label}_{category_slug}_{timestamp}.xlsx")
    save_excel(df, output_path)

    print("\n" + "=" * 55)
    print("  CRAWL SUMMARY")
    print("=" * 55)
    print(f"  Total machines scraped   : {len(df)}")
    print(f"  Excel saved at           : {os.path.abspath(output_path)}")
    print("-" * 55)
    print("  By type:")
    for t, cnt in df["Type"].value_counts().items():
        print(f"    {str(t)[:40]:<40}: {cnt}")
    print("=" * 55)
    print(f"\nDone. Open: {os.path.abspath(output_path)}")


def main():
    url = sys.argv[1] if len(sys.argv) > 1 else "https://dabrox.com/forging-stamping-machines/screw-presses/"
    scrape(url)


if __name__ == "__main__":
    main()
