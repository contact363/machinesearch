"""
Scraper for teyssou-ra.com/en_GB
- 14 categories, no pagination
- Extracts Type/Brand/Model/Description/Image from listing cards + detail pages
- No price (site uses Contact Us instead)
"""

import os
import re
import sys
import time
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from urllib.parse import urljoin

BASE_URL   = "https://www.teyssou-ra.com"
OUTPUT_DIR = "output"
NA         = "Not Available"
HEADERS    = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
    )
}

# All 14 categories — (Type label, listing URL)
ALL_CATEGORIES = [
    # Mechanical
    ("Centre d'Usinage",                   "https://www.teyssou-ra.com/en_GB/shop/category/27"),
    ("Fraiseuse",                           "https://www.teyssou-ra.com/en_GB/shop/category/32"),
    ("Rectifieuse",                         "https://www.teyssou-ra.com/en_GB/shop/category/39"),
    ("Scie",                                "https://www.teyssou-ra.com/en_GB/shop/category/42"),
    ("Tour",                                "https://www.teyssou-ra.com/en_GB/shop/category/47"),
    # Sheet Metal
    ("Cisaille Guillotine",                 "https://www.teyssou-ra.com/en_GB/shop/category/55"),
    ("Decoupe Laser",                       "https://www.teyssou-ra.com/en_GB/shop/category/57"),
    ("Ebavureuse",                          "https://www.teyssou-ra.com/en_GB/shop/category/58"),
    ("Encocheuse",                          "https://www.teyssou-ra.com/en_GB/shop/category/59"),
    ("Poinconneuse",                        "https://www.teyssou-ra.com/en_GB/shop/category/62"),
    ("Poinconneuse A Commande Numerique",   "https://www.teyssou-ra.com/en_GB/shop/category/63"),
    ("Presse Plieuse",                      "https://www.teyssou-ra.com/en_GB/shop/category/66"),
    ("Rouleuse",                            "https://www.teyssou-ra.com/en_GB/shop/category/67"),
    # Misc
    ("Materiel Divers",                     "https://www.teyssou-ra.com/en_GB/shop/category/69"),
]


def get_soup(url):
    for attempt in range(3):
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            r.raise_for_status()
            r.encoding = r.apparent_encoding or "utf-8"
            return BeautifulSoup(r.text, "lxml")
        except Exception as e:
            if attempt == 2:
                print(f"    [ERROR] {url} -> {e}")
            time.sleep(2)
    return None


def clean(text):
    if not text:
        return ""
    text = text.replace("\xa0", " ").replace("'", "'")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def full_image(src):
    """Upgrade image_512 to image_1024 for better resolution."""
    src = src.replace("/image_512/", "/image_1024/")
    return src if src.startswith("http") else urljoin(BASE_URL, src)


def collect_listing(cat_url, cat_type):
    """
    Scrape a category listing page.
    Product URL pattern: /en_GB/shop/[slug]-[id]
    Card structure (no CSS classes):
      <a href="product-url"><img src="..."></a>   ← image
      <a href="product-url">Title text</a>         ← title / description
      <div>Brand</div>                              ← brand
    """
    soup = get_soup(cat_url)
    if not soup:
        return []

    PRODUCT_RE = re.compile(r"/en_GB/shop/[^/?]+-\d+")

    # Gather all product URLs
    seen_urls = set()
    ordered   = []
    for a in soup.find_all("a", href=True):
        href = a["href"].split("?")[0]
        if PRODUCT_RE.search(href):
            full = href if href.startswith("http") else urljoin(BASE_URL, href)
            if full not in seen_urls:
                seen_urls.add(full)
                ordered.append(full)

    machines = []
    for product_url in ordered:
        # Find all <a> tags pointing to this product
        path    = product_url.replace(BASE_URL, "")
        anchors = soup.find_all("a", href=re.compile(re.escape(path)))

        description = NA
        image_url   = ""
        brand       = NA

        for a in anchors:
            img = a.find("img", src=True)
            if img:
                image_url = full_image(img["src"])
            else:
                txt = clean(a.get_text())
                if txt:
                    description = txt
                    # Brand is usually in the next sibling div
                    nxt = a.find_next_sibling()
                    if nxt and nxt.name in ("div", "span", "p"):
                        b = clean(nxt.get_text())
                        if b and len(b) < 50:
                            brand = b

        # Model: remove brand from description if brand is known
        model = NA
        if description != NA and brand != NA:
            # Try to find brand in title and take what comes after
            idx = description.upper().find(brand.upper())
            if idx != -1:
                after = description[idx + len(brand):].strip(" -–—:")
                if after:
                    model = after
            if model == NA or not model:
                model = description  # fallback: use full title

        machines.append({
            "Type":        cat_type,
            "Brand":       brand,
            "Model":       model,
            "Description": description,
            "Price":       "",
            "Image URL":   image_url,
            "Source URL":  product_url,
        })

    return machines


def parse_detail(url):
    """
    Fetch detail page for accurate Brand and Model.

    Brand: <span>Brand: <strong>HURON</strong></span>  — look for 'Brand:' label
    Model: parsed from H1 after removing the known brand name
    H1 patterns:
      "FRAISEUSE HURON MU66"            → brand=HURON  model=MU66
      "Cisaille guillotine - GPS 840 - AMADA" → brand=AMADA  model=GPS 840
      "Tour CNC SOMAB - UNIMAB 500"     → brand=SOMAB  model=UNIMAB 500
    """
    soup = get_soup(url)
    if not soup:
        return {}

    result = {}

    # ── Brand: look for "Brand:" / "Marque:" label near a <strong> ───────────
    brand = NA
    for elem in soup.find_all(["span", "div", "p", "li"]):
        txt = elem.get_text()
        if re.search(r"brand\s*:", txt, re.IGNORECASE) or re.search(r"marque\s*:", txt, re.IGNORECASE):
            strong = elem.find("strong")
            if strong:
                b = clean(strong.get_text())
                if b and b.lower() not in ("x", "-", ""):
                    brand = b
                    break

    if brand != NA:
        result["brand"] = brand

    # ── H1 = Description, parse Model from it ─────────────────────────────────
    h1 = soup.find("h1")
    if h1:
        h1_text = clean(h1.get_text())
        result["description"] = h1_text

        if brand != NA:
            # Pattern A: "Type - Model - Brand"  e.g. "Cisaille guillotine - GPS 840 - AMADA"
            parts = [p.strip() for p in h1_text.split(" - ")]
            model_found = False
            for i, part in enumerate(parts):
                if brand.upper() in part.upper():
                    # Brand is in this part — model is the OTHER parts (excluding Type first part)
                    other = [parts[j] for j in range(len(parts)) if j != i and j != 0]
                    if other:
                        result["model"] = " - ".join(other)
                        model_found = True
                        break

            # Pattern B: "Type Brand - Model"  e.g. "Tour CNC SOMAB - UNIMAB 500"
            if not model_found and " - " in h1_text:
                idx = h1_text.upper().find(brand.upper())
                if idx != -1:
                    after_brand = h1_text[idx + len(brand):].strip(" -–—:")
                    if after_brand:
                        result["model"] = after_brand
                        model_found = True

            # Pattern C: "TYPE BRAND MODEL"  e.g. "FRAISEUSE HURON MU66"
            if not model_found:
                idx = h1_text.upper().find(brand.upper())
                if idx != -1:
                    after_brand = h1_text[idx + len(brand):].strip()
                    if after_brand:
                        result["model"] = after_brand

    # ── Image (higher res) ────────────────────────────────────────────────────
    for img in soup.find_all("img", src=True):
        src = img["src"]
        if "product.product" in src or "product.template" in src:
            result["image_url"] = src if src.startswith("http") else urljoin(BASE_URL, src)
            break

    return result


def save_excel(df, filepath):
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    df.to_excel(filepath, index=False, engine="openpyxl")

    wb = load_workbook(filepath)
    ws = wb.active

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    ws.freeze_panes = "A2"

    for col_idx, col_cells in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len    = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 80)

    wb.save(filepath)


def main():
    print("Scraper — teyssou-ra.com")
    print("=" * 60)
    print(f"  Categories: {len(ALL_CATEGORIES)}")
    print("=" * 60)

    # Step 1: Collect all machine URLs from listing pages
    print("\nStep 1: Collecting machines from all category pages...")
    all_machines = []
    seen_urls    = set()

    for cat_type, cat_url in ALL_CATEGORIES:
        machines = collect_listing(cat_url, cat_type)
        new = 0
        for m in machines:
            if m["Source URL"] not in seen_urls:
                seen_urls.add(m["Source URL"])
                all_machines.append(m)
                new += 1
        print(f"  {cat_type:<40}: {new} machines")
        time.sleep(0.5)

    print(f"\n  Total unique machines: {len(all_machines)}")

    # Step 2: Fetch detail pages for accurate Brand / Model
    print("\nStep 2: Fetching detail pages for Brand & Model...")
    for i, m in enumerate(all_machines, 1):
        print(f"  [{i}/{len(all_machines)}] {m['Source URL']}")
        detail = parse_detail(m["Source URL"])

        if detail.get("brand") and detail["brand"] != NA:
            m["Brand"] = detail["brand"]
        if detail.get("model") and detail["model"] != NA:
            m["Model"] = detail["model"]
        if detail.get("description"):
            m["Description"] = detail["description"]
        if detail.get("image_url"):
            m["Image URL"] = detail["image_url"]

        time.sleep(0.5)

    if not all_machines:
        print("No data collected.")
        return

    df = pd.DataFrame(all_machines)
    df = df[["Type", "Brand", "Model", "Description", "Price", "Image URL", "Source URL"]]
    df["Brand"]       = df["Brand"].fillna(NA)
    df["Model"]       = df["Model"].fillna(NA)
    df["Description"] = df["Description"].fillna(NA)
    df["Price"]       = df["Price"].fillna("")
    df["Image URL"]   = df["Image URL"].fillna("")

    df.sort_values("Type", inplace=True)
    df.reset_index(drop=True, inplace=True)

    timestamp   = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = os.path.join(OUTPUT_DIR, f"teyssou_output_{timestamp}.xlsx")
    save_excel(df, output_path)

    print("\n" + "=" * 60)
    print("  CRAWL SUMMARY")
    print("=" * 60)
    print(f"  Total machines   : {len(df)}")
    print(f"  Excel saved at   : {os.path.abspath(output_path)}")
    print("-" * 60)
    print("  By Type:")
    for t, cnt in df["Type"].value_counts().items():
        print(f"    {str(t):<40}: {cnt}")
    print("=" * 60)
    print(f"\nDone. Open: {os.path.abspath(output_path)}")


if __name__ == "__main__":
    main()
