import os
import re
import time
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
BASE_URL   = "https://valmatrading.com"
OUTPUT_DIR = "output"
NA         = "Not Available"
HEADERS    = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
    )
}

CATEGORIES = {
    "Mechanical C-Frame Presses":        "mechanical-c-frame-presses",
    "Mechanical Double Column Presses":   "mechanical-double-column-presses",
    "Hydraulic C-Frame Presses":         "hydraulic-c-frame-presses",
    "Hydraulic Double Column Presses":   "hydraulic-double-column-presses",
    "Fine Blanking Presses":             "fine-blanking-presses",
    "Hot Forging Presses":               "hot-forging-presses",
    "Screw Fly Presses":                 "screw-fly-presses",
    "CNC Machining":                     "cnc-machining",
    "Servo Press Equipment":             "servo-press-equipment",
    "Lathes":                            "lathes",
    "Milling Machinery":                 "milling-machinery",
    "Others":                            "others",
}
# ─────────────────────────────────────────────


def get_soup(url):
    for attempt in range(3):
        try:
            r = requests.get(url, headers=HEADERS, timeout=15)
            r.raise_for_status()
            r.encoding = "utf-8"
            return BeautifulSoup(r.text, "lxml")
        except Exception as e:
            if attempt == 2:
                print(f"    [ERROR] {url} -> {e}")
            time.sleep(2)
    return None


def clean(text):
    if not text:
        return ""
    text = text.replace("\xa0", " ").replace("’", "'")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def full_image_url(src):
    """Remove WordPress thumbnail size suffix e.g. -300x225 before extension."""
    src = re.sub(r"-\d+x\d+(\.(jpg|jpeg|png|webp))$", r"\1", src, flags=re.IGNORECASE)
    return src


def collect_product_urls(category_slug, type_label):
    url = f"{BASE_URL}/{category_slug}/"
    soup = get_soup(url)
    if not soup:
        return []

    seen = set()
    products = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if "/products/" in href and href not in seen:
            seen.add(href)
            products.append({"url": href, "type": type_label})

    return products


def parse_machine(url, machine_type):
    soup = get_soup(url)
    if not soup:
        return None

    # Remove nav/footer/header noise
    for tag in soup(["script", "style", "nav", "footer", "header"]):
        tag.decompose()

    main = soup.find("main") or soup.find("article") or soup.body

    # ── Title / Description ───────────────────
    # Try gb-headline-text span first, then page <title>
    headline = soup.find("span", class_="gb-headline-text")
    if headline:
        description = clean(headline.get_text())
    else:
        title_tag = soup.find("title")
        raw_title = clean(title_tag.get_text()) if title_tag else ""
        description = raw_title.split("|")[0].strip() if raw_title else NA

    # ── Extract label:value pairs from main content ──
    # Get all text lines from main
    raw_lines = [clean(l) for l in main.get_text(separator="\n").splitlines() if clean(l)]

    def find_value(labels):
        """Find value for given label keywords in the line list."""
        for i, line in enumerate(raw_lines):
            low = line.lower().rstrip(":")
            for lbl in labels:
                if low == lbl.lower():
                    # Value is on next non-empty line
                    for j in range(i + 1, min(i + 4, len(raw_lines))):
                        candidate = raw_lines[j]
                        if candidate and candidate.lower() not in [l.lower() for l in labels]:
                            return candidate
            # Also handle "Label: Value" on same line
            for lbl in labels:
                if line.lower().startswith(lbl.lower() + ":"):
                    val = line[len(lbl) + 1:].strip()
                    if val:
                        return val
        return NA

    # ── Brand ─────────────────────────────────
    brand = find_value(["Brand", "Manufacturer", "Make"])
    if brand == NA and description != NA:
        # Fallback: first word of title
        words = description.split()
        if words:
            brand = words[0].strip(",:;-()")

    # ── Model ─────────────────────────────────
    model = find_value(["Model", "Model number", "Type"])

    # ── Price ─────────────────────────────────
    price = find_value(["Price", "Asking price", "Selling price"])
    if price == NA:
        price = ""  # Leave blank — not on this website

    # ── Image URL ─────────────────────────────
    image_url = ""
    for img in soup.find_all("img", src=True):
        src = img["src"]
        if "wp-content/uploads" in src and "logo" not in src.lower():
            image_url = full_image_url(src)
            break

    # ── Log missing ───────────────────────────
    missing = [f for f, v in {"Brand": brand, "Model": model, "Description": description}.items() if v == NA]
    if missing:
        print(f"    [MISSING] {', '.join(missing)}")

    return {
        "Type":        machine_type,
        "Brand":       brand,
        "Model":       model,
        "Description": description,
        "Price":       price,
        "Image URL":   image_url,
        "Source URL":  url,
    }


def save_excel(df, filepath):
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    df.to_excel(filepath, index=False, engine="openpyxl")

    wb = load_workbook(filepath)
    ws = wb.active

    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    ws.freeze_panes = "A2"

    for col_idx, col_cells in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len    = max((len(str(c.value)) for c in col_cells if c.value), default=10)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 80)

    wb.save(filepath)


def main():
    print("Scraper — valmatrading.com")
    print("=" * 55)

    # Step 1: Collect all product URLs
    print("Step 1: Collecting product URLs from all categories...")
    all_refs  = []
    seen_urls = set()

    for type_label, slug in CATEGORIES.items():
        refs = collect_product_urls(slug, type_label)
        count = 0
        for ref in refs:
            if ref["url"] not in seen_urls:
                seen_urls.add(ref["url"])
                all_refs.append(ref)
                count += 1
        print(f"  {type_label:<35}: {count} machines")
        time.sleep(0.5)

    print(f"\n  Total unique machines: {len(all_refs)}")

    # Step 2: Scrape each machine
    print("\nStep 2: Scraping machine detail pages...")
    all_machines = []
    missing_stats = {"Brand": 0, "Model": 0, "Description": 0}

    for i, ref in enumerate(all_refs, 1):
        print(f"  [{i}/{len(all_refs)}] {ref['url']}")
        try:
            data = parse_machine(ref["url"], ref["type"])
            if data:
                for field in missing_stats:
                    if data.get(field) == NA:
                        missing_stats[field] += 1
                all_machines.append(data)
        except Exception as e:
            print(f"    [SKIP] {e}")
        time.sleep(0.5)

    if not all_machines:
        print("No data collected.")
        return

    df = pd.DataFrame(all_machines)
    total_raw = len(df)
    df.drop_duplicates(subset=["Source URL"], inplace=True)
    dupes = total_raw - len(df)

    df["Brand"].fillna(NA, inplace=True)
    df["Model"].fillna(NA, inplace=True)
    df["Description"].fillna(NA, inplace=True)
    df["Price"].fillna("", inplace=True)
    df["Image URL"].fillna("", inplace=True)

    df.sort_values("Type", inplace=True)
    df.reset_index(drop=True, inplace=True)

    timestamp   = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = os.path.join(OUTPUT_DIR, f"valma_output_{timestamp}.xlsx")
    save_excel(df, output_path)

    print("\n" + "=" * 55)
    print("  CRAWL SUMMARY")
    print("=" * 55)
    print(f"  Total machines scraped   : {total_raw}")
    print(f"  Duplicates removed       : {dupes}")
    print(f"  Final machines in Excel  : {len(df)}")
    print(f"  Excel saved at           : {os.path.abspath(output_path)}")
    print("-" * 55)
    print("  By machine type:")
    for t, cnt in df["Type"].value_counts().items():
        t_str = str(t)[:35]
        print(f"    {t_str:<35}: {cnt}")
    print("-" * 55)
    print("  Missing field statistics:")
    for field, count in missing_stats.items():
        pct = (count / total_raw * 100) if total_raw > 0 else 0
        print(f"    {field:<15}: {count:>3} missing ({pct:.1f}%)")
    print("=" * 55)
    print(f"\nDone. Open: {os.path.abspath(output_path)}")


if __name__ == "__main__":
    main()
